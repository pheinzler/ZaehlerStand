#!/usr/bin/python3

from pathlib import Path
import openpyxl as xls

def main():
    wb = xls.load_workbook(filename='Python_xls_test.xlsx')
    abc = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    ws = wb.create_sheet(title='NewWS_test')

    for row in range(24):
        for col in abc:
            coord = col + str(row)
            ws[coord] = coord
            print(coord)
    wb.save('Python_xls_test.xlsx')

if '__name__' == __name__:
    main()
